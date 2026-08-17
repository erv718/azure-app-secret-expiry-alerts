"""End-to-end tests for monitor.py using a local mock Slack webhook server.

Needs no Azure tenant and no Slack workspace: Microsoft Graph calls are
stubbed and Slack webhooks point at a local HTTP server. Run with:

    python tests/test_monitor.py

Exits nonzero if any check fails.
"""
import json
import os
import sys
import tempfile
import threading
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path

PROJECT = Path(__file__).resolve().parents[1]
WORKDIR = Path(tempfile.mkdtemp(prefix="monitor_test_"))

# --- Mock Slack webhook server (port chosen by the OS) -----------------------
received = []          # list of (path, payload)
fail_paths = set()     # paths that should return HTTP 500


class Handler(BaseHTTPRequestHandler):
    def do_POST(self):
        body = self.rfile.read(int(self.headers.get("Content-Length", 0)))
        payload = json.loads(body)
        if self.path in fail_paths:
            self.send_response(500)
            self.end_headers()
            return
        received.append((self.path, payload))
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"ok")

    def log_message(self, *args):
        pass


server = HTTPServer(("127.0.0.1", 0), Handler)
PORT = server.server_address[1]
threading.Thread(target=server.serve_forever, daemon=True).start()

# --- Safe env BEFORE importing monitor (load_dotenv won't override these) ----
SAFE_ENV = {
    "AZURE_AUTH_METHOD": "service_principal",
    "AZURE_TENANT_ID": "00000000-0000-0000-0000-000000000000",
    "AZURE_CLIENT_ID": "00000000-0000-0000-0000-000000000000",
    "AZURE_CLIENT_SECRET": "fake",
    "ALERT_THRESHOLD_DAYS": "90,60,30,14,7",
    "FILTER_INCLUDE_APP_IDS": "", "FILTER_INCLUDE_NAMES": "",
    "FILTER_EXCLUDE_APP_IDS": "", "FILTER_EXCLUDE_NAMES": "",
    "SLACK_ENABLED": "true",
    "SLACK_WEBHOOK_URL": f"http://127.0.0.1:{PORT}/hook/default",
    "SLACK_BOT_TOKEN": "",
    "SLACK_CHANNEL_ID": "",
    "EMAIL_ENABLED": "false",
    "SMTP_HOST": "", "SMTP_PORT": "587", "SMTP_USERNAME": "",
    "SMTP_PASSWORD": "", "SMTP_USE_TLS": "true", "EMAIL_FROM": "", "EMAIL_TO": "",
    "STATE_FILE_PATH": str(WORKDIR / "test_state.json"),
}
os.environ.update(SAFE_ENV)

sys.path.insert(0, str(PROJECT))
import monitor  # noqa: E402

monitor.CHANNELS_FILE = WORKDIR / "test_channels.json"
monitor.CHANNELS_FILE.write_text(json.dumps({
    "teams": [{
        "name": "D365 Team",
        "slack_webhook_url": f"http://127.0.0.1:{PORT}/hook/d365",
        "slack_channel_id": "",
        "app_ids": ["aaaaaaaa-1111-2222-3333-444444444444"],
        "app_name_patterns": ["d365", "recurring integrations", "spsedi", "3pl integration"],
    }]
}), encoding="utf-8")

# --- Fake Graph data ---------------------------------------------------------

def iso(days):
    return (datetime.now(timezone.utc) + timedelta(days=days)).strftime("%Y-%m-%dT%H:%M:%SZ")


def fake_apps_v1():
    return [
        {"displayName": "D365 Data Lake PROD", "appId": "15c4612c-a719-47c0-80e2-5207067fcabc",
         "passwordCredentials": [{"displayName": "prod-secret", "keyId": "k1", "endDateTime": iso(5)}],
         "keyCredentials": []},
        {"displayName": "Recurring Integrations App", "appId": "e56b34fe-76de-47e4-bfba-d7040b00d621",
         "passwordCredentials": [{"displayName": "ris", "keyId": "k2", "endDateTime": iso(25)}],
         "keyCredentials": []},
        {"displayName": "SPSEDI Application", "appId": "b149b482-1499-4e06-b880-91ea1c1c4efc",
         "passwordCredentials": [], "keyCredentials": [{"displayName": "spsedi-cert", "keyId": "k3", "endDateTime": iso(-3)}]},
        {"displayName": "RouteByIdApp", "appId": "aaaaaaaa-1111-2222-3333-444444444444",
         "passwordCredentials": [{"displayName": "byid", "keyId": "k4", "endDateTime": iso(60)}],
         "keyCredentials": []},
        # Non-team apps (default channel only)
        {"displayName": "=HYPERLINK(\"http://evil\",\"click\")", "appId": "dddddddd-0000-0000-0000-000000000001",
         "passwordCredentials": [{"displayName": "inj", "keyId": "k5", "endDateTime": iso(10)}],
         "keyCredentials": []},
        {"displayName": "Payments <!channel> & Co \x07Ctrl", "appId": "dddddddd-0000-0000-0000-000000000002",
         "passwordCredentials": [{"displayName": None, "keyId": "k6", "endDateTime": iso(2)}],
         "keyCredentials": []},
        {"displayName": "Healthy App", "appId": "dddddddd-0000-0000-0000-000000000003",
         "passwordCredentials": [{"displayName": "fine", "keyId": "k7", "endDateTime": iso(4000)}],
         "keyCredentials": []},
    ]


current_apps = fake_apps_v1()
monitor.get_access_token = lambda: "fake-token"
monitor.get_applications = lambda token: current_apps

# --- Helpers -----------------------------------------------------------------
passed, failed = [], []


def check(name, cond, detail=""):
    (passed if cond else failed).append(name)
    print(("PASS " if cond else "FAIL ") + name + (f"  [{detail}]" if detail and not cond else ""))


def reset_received():
    del received[:]


def paths():
    return [p for p, _ in received]


def payload_for(path):
    return [pl for p, pl in received if p == path]


def all_section_texts(payload):
    out = []
    for b in payload.get("blocks", []):
        if b.get("type") == "section":
            out.append(b["text"]["text"])
        if b.get("type") == "context":
            out.extend(e["text"] for e in b.get("elements", []))
    return out


def state():
    return json.loads(monitor.STATE_FILE.read_text(encoding="utf-8"))


# ============================================================================
print("\n--- Run 1: first run, both channels should get a report ---")
reset_received()
rc = monitor.main([])
check("run1 exit code 0", rc == 0)
check("run1 default posted", "/hook/default" in paths())
check("run1 d365 posted", "/hook/d365" in paths())

d365_payload = payload_for("/hook/d365")[0]
d365_text = json.dumps(d365_payload)
check("d365 gets RouteByIdApp (app_ids match)", "RouteByIdApp" in d365_text)
check("d365 gets pattern matches", "D365 Data Lake PROD" in d365_text and "SPSEDI" in d365_text)
check("d365 does NOT get default-only apps", "Payments" not in d365_text and "HYPERLINK" not in d365_text)

default_payload = payload_for("/hook/default")[0]
check("default payload has fallback text", bool(default_payload.get("text")))
check("default mrkdwn-escapes <!channel>", "<!channel>" not in json.dumps(default_payload)
      and "&lt;!channel&gt;" in json.dumps(default_payload))
for name, pl in (("default", default_payload), ("d365", d365_payload)):
    check(f"{name} block count <= 50", len(pl["blocks"]) <= 50, str(len(pl["blocks"])))
    check(f"{name} sections <= 3000 chars", all(len(t) <= 3000 for t in all_section_texts(pl)))

st = state()
check("state has both channels", set(st["channels"]) == {"default", "D365 Team"}, str(set(st["channels"])))
check("state v2 format", st.get("version") == 2)

# ============================================================================
print("\n--- Run 2: no changes -> nothing sent ---")
reset_received()
monitor.main([])
check("run2 nothing posted", not received, str(paths()))

# ============================================================================
print("\n--- Run 3: new team credential, but team webhook fails -> retry state kept ---")
current_apps.append(
    {"displayName": "3PL Integration Prod", "appId": "a29c59d4-f9d2-4a7e-ba8c-d24351db8b3c",
     "passwordCredentials": [{"displayName": "bergen", "keyId": "k8", "endDateTime": iso(12)}],
     "keyCredentials": []})
reset_received()
fail_paths.add("/hook/d365")
monitor.main([])
check("run3 default posted", "/hook/default" in paths())
check("run3 d365 delivery failed (nothing recorded)", "/hook/d365" not in paths())
st = state()
check("run3 d365 state slice NOT updated (no 3PL key)",
      not any("a29c59d4" in k for k in st["channels"]["D365 Team"]["credential_keys"]))
check("run3 default state updated (has 3PL key)",
      any("a29c59d4" in k for k in st["channels"]["default"]["credential_keys"]))

print("\n--- Run 4: team webhook healthy again -> team retries, default suppressed ---")
fail_paths.discard("/hook/d365")
reset_received()
monitor.main([])
check("run4 d365 retried", "/hook/d365" in paths())
check("run4 default suppressed", "/hook/default" not in paths())
d365_retry = payload_for("/hook/d365")[0]
check("run4 d365 message includes new 3PL app", "3PL Integration Prod" in json.dumps(d365_retry))

# ============================================================================
print("\n--- Run 5: everything resolved -> all-clear to both channels ---")
del current_apps[:]
current_apps.append({"displayName": "Healthy App", "appId": "dddddddd-0000-0000-0000-000000000003",
                     "passwordCredentials": [{"displayName": "fine", "keyId": "k7", "endDateTime": iso(4000)}],
                     "keyCredentials": []})
reset_received()
monitor.main([])
check("run5 default all-clear", any("All clear" in json.dumps(pl) for pl in payload_for("/hook/default")))
check("run5 d365 all-clear", any("All clear" in json.dumps(pl) for pl in payload_for("/hook/d365")))
st = state()
check("run5 state cleared", not st["channels"]["default"]["credential_keys"]
      and not st["channels"]["D365 Team"]["credential_keys"])

print("\n--- Run 6: still empty -> silent ---")
reset_received()
monitor.main([])
check("run6 nothing posted", not received, str(paths()))

# ============================================================================
print("\n--- Dry run: prints routing, sends nothing, saves nothing ---")
current_apps[:] = fake_apps_v1()
reset_received()
st_before = state()
rc = monitor.main(["--dry-run"])
check("dry-run exit 0", rc == 0)
check("dry-run nothing posted", not received)
check("dry-run state untouched", state() == st_before)

# ============================================================================
print("\n--- --test-channels: test message to every channel ---")
reset_received()
rc = monitor.main(["--test-channels"])
check("test-channels exit 0", rc == 0)
check("test-channels hit default", "/hook/default" in paths())
check("test-channels hit d365", "/hook/d365" in paths())
check("test-channels message says test", all("test" in json.dumps(pl).lower() for _, pl in received))

# ============================================================================
print("\n--- Excel: formula injection + illegal chars neutralized ---")
from openpyxl import load_workbook  # noqa: E402
alerts = monitor.check_expiry(fake_apps_v1(), [90, 60, 30, 14, 7])
xlsx = monitor._generate_excel(alerts)
wb = load_workbook(xlsx)
ws = wb.active
formula_cells = [c.coordinate for row in ws.iter_rows() for c in row if c.data_type == "f"]
check("no formula cells in workbook", not formula_cells, str(formula_cells))
names = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
check("injected name kept as literal text", any(n and n.startswith("=HYPERLINK") for n in names))
check("control char stripped", all("\x07" not in (n or "") for n in names))
os.unlink(xlsx)

# ============================================================================
print("\n--- v1 state file migration ---")
monitor.STATE_FILE.write_text(json.dumps({
    "last_run": "2026-01-01T00:00:00+00:00",
    "buckets": {"expired": 1, "7_days": 0},
    "credential_keys": ["x|Secret|old"],
}), encoding="utf-8")
migrated = monitor._load_state()
check("v1 migrates to default channel", "default" in migrated
      and migrated["default"]["credential_keys"] == ["x|Secret|old"])

# ============================================================================
print("\n--- Slack 3000-char limit under huge new-credential lists ---")
huge = [{"app_name": f"App With A Fairly Long Name Number {i:04d}", "app_id": f"{i:032d}",
         "credential_type": "Secret", "credential_name": f"secret-{i}", "key_id": f"kk{i}",
         "created": "2025-08-01", "expires": "2026-08-01", "days_left": 6} for i in range(2000)]
reset_received()
buckets = monitor._bucket_counts(huge, [90, 60, 30, 14, 7])
ok = monitor.send_slack_alert(f"http://127.0.0.1:{PORT}/hook/default", "Load test", huge,
                              buckets, ["First run - no previous state"], huge,
                              [90, 60, 30, 14, 7], None, "", "")
check("huge alert delivered", ok)
pl = payload_for("/hook/default")[0]
check("huge alert <= 50 blocks", len(pl["blocks"]) <= 50, str(len(pl["blocks"])))
check("huge alert sections <= 3000", all(len(t) <= 3000 for t in all_section_texts(pl)))
check("huge alert notes truncation", "truncated" in json.dumps(pl).lower())

# ============================================================================
print("\n--- Config validation: malformed channels.json entries ---")
GOOD_CHANNELS = monitor.CHANNELS_FILE.read_text(encoding="utf-8")
monitor.CHANNELS_FILE.write_text(json.dumps({"teams": [
    {"name": "D365 Team", "slack_webhook_url": f"http://127.0.0.1:{PORT}/hook/d365",
     "app_name_patterns": ["d365"]},
    {"name": "NullIds", "slack_webhook_url": "http://x/", "app_ids": None,
     "app_name_patterns": ["zzz-no-match"]},
    {"name": "StrPatterns", "slack_webhook_url": "http://x/", "app_name_patterns": "prod"},
    {"name": "Dup", "slack_webhook_url": "http://first/", "app_ids": ["x"]},
    {"name": "Dup", "slack_webhook_url": "http://second/", "app_ids": ["y"]},
]}), encoding="utf-8")
teams = monitor.load_channels()
team_names = [t["name"] for t in teams]
check("null app_ids tolerated as empty", "NullIds" in team_names
      and teams[team_names.index("NullIds")]["app_ids"] == [])
check("string patterns team skipped", "StrPatterns" not in team_names)
check("duplicate team keeps first only", team_names.count("Dup") == 1
      and teams[team_names.index("Dup")]["slack_webhook_url"] == "http://first/")
try:
    routed = monitor.route_alerts(monitor.check_expiry(fake_apps_v1(), [90, 60, 30, 14, 7]), teams)
    check("route_alerts survives malformed-adjacent config", True)
    check("no char-level pattern flooding", not routed.get("NullIds"))
except TypeError as e:
    check("route_alerts survives malformed-adjacent config", False, str(e))
monitor.CHANNELS_FILE.write_text(GOOD_CHANNELS, encoding="utf-8")

# ============================================================================
print("\n--- v1 legacy-key migration: no false 'new credentials' storm ---")
monitor.STATE_FILE.unlink()
current_apps[:] = [
    {"displayName": "Legacy App", "appId": "11111111-1111-1111-1111-111111111111",
     "passwordCredentials": [{"displayName": "legacy-secret", "keyId": "kx", "endDateTime": iso(5)}],
     "keyCredentials": []},
]
monitor.STATE_FILE.write_text(json.dumps({
    "last_run": "2026-01-01T00:00:00+00:00",
    "buckets": {"expired": 0, "7_days": 1, "14_days": 0, "30_days": 0, "90_days": 0, "90_plus": 0},
    "credential_keys": ["11111111-1111-1111-1111-111111111111|Secret|legacy-secret"],
}), encoding="utf-8")
reset_received()
rc = monitor.main([])
check("v1 upgrade run silent (no storm)", not received, str(paths()))
check("v1 upgrade exit 0", rc == 0)

print("\n--- Bucket transition without new creds lists 'Currently flagged' ---")
current_apps[0]["passwordCredentials"][0]["endDateTime"] = iso(-2)  # same keyId, now expired
reset_received()
rc = monitor.main([])
check("transition run posted", "/hook/default" in paths())
pl = payload_for("/hook/default")[0]
check("transition lists the credential", "Legacy App" in json.dumps(pl))
check("transition uses 'Currently flagged' heading", "Currently flagged" in json.dumps(pl))
check("transition includes credential name", "legacy-secret" in json.dumps(pl))

print("\n--- Transport change forces a rebaseline send ---")
os.environ["EMAIL_ENABLED"] = "true"   # SMTP_HOST empty -> unconfigured, excluded
reset_received()
rc = monitor.main([])
check("transport change triggers send", "/hook/default" in paths())
check("transport change noted in message",
      "configuration changed" in json.dumps(payload_for("/hook/default")[0]))
check("unconfigured email excluded from gate (exit 0)", rc == 0)
reset_received()
rc = monitor.main([])
check("next run silent (state saved despite email misconfig)", not received, str(paths()))
os.environ["EMAIL_ENABLED"] = "false"

print("\n--- Unreadable channels.json: exit 1, team state preserved ---")
current_apps.append(
    {"displayName": "Brand New App", "appId": "22222222-2222-2222-2222-222222222222",
     "passwordCredentials": [{"displayName": "bn", "keyId": "kz", "endDateTime": iso(3)}],
     "keyCredentials": []})
st_channels_before = set(state()["channels"])
monitor.CHANNELS_FILE.write_text("not json {{{", encoding="utf-8")
reset_received()
rc = monitor.main([])
check("unreadable config exit 1", rc == 1)
check("default still notified", "/hook/default" in paths())
check("team state not pruned on unreadable config",
      "D365 Team" not in st_channels_before or "D365 Team" in set(state()["channels"]))
monitor.CHANNELS_FILE.write_text(GOOD_CHANNELS, encoding="utf-8")

print("\n--- Delivery failure: exit 1 and state kept for retry ---")
current_apps.append(
    {"displayName": "Another App", "appId": "33333333-3333-3333-3333-333333333333",
     "passwordCredentials": [{"displayName": "aa", "keyId": "ka", "endDateTime": iso(2)}],
     "keyCredentials": []})
fail_paths.add("/hook/default")
reset_received()
rc = monitor.main([])
check("failed delivery exit 1", rc == 1)
check("failed delivery does not record new key",
      not any("33333333" in k for k in state()["channels"]["default"]["credential_keys"]))
fail_paths.discard("/hook/default")

print("\n--- Graph 7-digit fractional seconds parse (Python < 3.11 compat) ---")
parsed = monitor._parse_date("2027-01-01T00:00:00.1234567Z")
check("fractional-second date parses", parsed is not None)
check("plain date still parses", monitor._parse_date("2027-01-01T00:00:00Z") is not None)

# ============================================================================
server.shutdown()
print(f"\n{'=' * 60}\nRESULT: {len(passed)} passed, {len(failed)} failed")
if failed:
    print("FAILED:", *failed, sep="\n  ")
sys.exit(1 if failed else 0)
