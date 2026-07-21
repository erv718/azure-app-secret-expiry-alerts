#!/usr/bin/env python3
"""Azure App Registration Secret & Certificate Expiry Monitor.

Queries Microsoft Graph API for all app registrations, checks secrets and
certificates against configurable expiry thresholds, and sends alerts via
Slack and/or email. Optional per-team routing sends scoped alerts to
additional Slack channels (see channels.template.json).
"""

import argparse
import json
import logging
import os
import re
import smtplib
import sys
import tempfile
from datetime import datetime, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import requests
from azure.identity import ClientSecretCredential, ManagedIdentityCredential
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger(__name__)

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
# Override for deployments without a persistent working directory (containers,
# serverless) - point it at mounted/persistent storage.
STATE_FILE = Path(os.getenv("STATE_FILE_PATH", "") or Path(__file__).parent / ".last_run_state.json")
CHANNELS_FILE = Path(__file__).parent / "channels.json"

# The env-var-configured Slack webhook + email recipients act as the default
# catch-all channel and always receive the full tenant-wide report.
DEFAULT_CHANNEL = "default"

# Slack hard limits: 50 blocks per message, 3000 chars per section text.
MAX_BLOCKS = 45
SECTION_CHAR_BUDGET = 2900

# Characters openpyxl refuses to write into a cell.
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

# Graph emits 7-digit fractional seconds; fromisoformat before Python 3.11
# only accepts exactly 3 or 6 digits.
_FRACTION_RE = re.compile(r"\.(\d{1,6})\d*")


# ---------------------------------------------------------------------------
# Azure auth
# ---------------------------------------------------------------------------

def get_access_token() -> str:
    method = os.getenv("AZURE_AUTH_METHOD", "service_principal").lower()
    scope = "https://graph.microsoft.com/.default"

    if method == "managed_identity":
        credential = ManagedIdentityCredential()
    else:
        tenant = os.environ["AZURE_TENANT_ID"]
        client_id = os.environ["AZURE_CLIENT_ID"]
        client_secret = os.environ["AZURE_CLIENT_SECRET"]
        credential = ClientSecretCredential(tenant, client_id, client_secret)

    token = credential.get_token(scope)
    return token.token


# ---------------------------------------------------------------------------
# Microsoft Graph queries
# ---------------------------------------------------------------------------

def get_applications(token: str) -> list[dict]:
    """Fetch all app registrations (handles paging)."""
    headers = {"Authorization": f"Bearer {token}"}
    url = f"{GRAPH_BASE}/applications?$select=id,displayName,appId,passwordCredentials,keyCredentials&$top=999"
    apps = []

    while url:
        resp = requests.get(url, headers=headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        apps.extend(data.get("value", []))
        url = data.get("@odata.nextLink")

    return apps


# ---------------------------------------------------------------------------
# App filtering
# ---------------------------------------------------------------------------

def filter_applications(apps: list[dict]) -> list[dict]:
    """Apply optional include/exclude filters from config."""
    include_ids = _csv_set("FILTER_INCLUDE_APP_IDS")
    include_names = _csv_list("FILTER_INCLUDE_NAMES")
    exclude_ids = _csv_set("FILTER_EXCLUDE_APP_IDS")
    exclude_names = _csv_list("FILTER_EXCLUDE_NAMES")

    filtered = apps

    if include_ids:
        filtered = [a for a in filtered if a.get("appId", "") in include_ids]

    if include_names:
        filtered = [
            a for a in filtered
            if any(p in a.get("displayName", "").lower() for p in include_names)
        ]

    if exclude_ids:
        filtered = [a for a in filtered if a.get("appId", "") not in exclude_ids]

    if exclude_names:
        filtered = [
            a for a in filtered
            if not any(p in a.get("displayName", "").lower() for p in exclude_names)
        ]

    if len(filtered) != len(apps):
        log.info("Filtered to %d app(s) (from %d total)", len(filtered), len(apps))

    return filtered


def _csv_set(env_var: str) -> set[str]:
    raw = os.getenv(env_var, "")
    return {v.strip() for v in raw.split(",") if v.strip()} if raw.strip() else set()


def _csv_list(env_var: str) -> list[str]:
    raw = os.getenv(env_var, "")
    return [v.strip().lower() for v in raw.split(",") if v.strip()] if raw.strip() else []


# ---------------------------------------------------------------------------
# Expiry checking
# ---------------------------------------------------------------------------

def parse_thresholds() -> list[int]:
    raw = os.getenv("ALERT_THRESHOLD_DAYS", "90,60,30,14,7")
    days = sorted({int(d.strip()) for d in raw.split(",") if d.strip()})
    if not days:
        raise ValueError("ALERT_THRESHOLD_DAYS contains no usable values")
    return days


def check_expiry(apps: list[dict], threshold_days: list[int]) -> list[dict]:
    """Return a list of expiring/expired credentials."""
    now = datetime.now(timezone.utc)
    max_threshold = max(threshold_days)
    alerts = []

    for app in apps:
        for kind, cred_list in (("Secret", "passwordCredentials"), ("Certificate", "keyCredentials")):
            for cred in app.get(cred_list, []):
                end = _parse_date(cred.get("endDateTime"))
                if end is None:
                    continue
                days_left = (end - now).days
                if days_left <= max_threshold:
                    start = _parse_date(cred.get("startDateTime"))
                    alerts.append({
                        "app_name": app.get("displayName", "Unknown"),
                        "app_id": app.get("appId", ""),
                        "credential_type": kind,
                        "credential_name": cred.get("displayName") or "(unnamed)",
                        "key_id": cred.get("keyId", ""),
                        "created": start.strftime("%Y-%m-%d") if start else "",
                        "expires": end.strftime("%Y-%m-%d"),
                        "days_left": days_left,
                    })

    alerts.sort(key=lambda a: a["days_left"])
    return alerts


def _parse_date(value: str | None) -> datetime | None:
    if not value:
        return None
    normalized = _FRACTION_RE.sub(lambda m: "." + m.group(1).ljust(6, "0"),
                                  value.replace("Z", "+00:00"))
    try:
        return datetime.fromisoformat(normalized)
    except (ValueError, TypeError):
        log.warning("Could not parse credential end date: %r", value)
        return None


# ---------------------------------------------------------------------------
# Team channel routing
# ---------------------------------------------------------------------------

def _validated_str_list(value, team_name: str, field: str) -> list[str] | None:
    """Return a cleaned list of strings, or None if the field is malformed."""
    if value is None:
        return []
    if isinstance(value, list) and all(isinstance(v, str) for v in value):
        return [v.strip() for v in value if v and v.strip()]
    log.warning("Team '%s': %s must be a list of strings - skipping this team", team_name, field)
    return None


def load_channels() -> list[dict] | None:
    """Load optional per-team routing config from channels.json.

    Returns [] when the file is absent, None when it exists but cannot be
    read/parsed (so callers can avoid destructive actions like state pruning).
    """
    if not CHANNELS_FILE.exists():
        return []
    try:
        data = json.loads(CHANNELS_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        log.error("Could not read %s: %s", CHANNELS_FILE.name, exc)
        return None

    teams = data.get("teams", []) if isinstance(data, dict) else []
    valid = []
    seen_names = set()
    for team in teams:
        if not isinstance(team, dict) or not team.get("name"):
            log.warning("Skipping channels.json entry with no name")
            continue
        name = team["name"]
        if name == DEFAULT_CHANNEL:
            log.warning("Team name '%s' is reserved, skipping", DEFAULT_CHANNEL)
            continue
        if name in seen_names:
            log.warning("Duplicate team name '%s' in channels.json - skipping the duplicate", name)
            continue
        if not team.get("slack_webhook_url"):
            log.warning("Team '%s' has no slack_webhook_url, skipping", name)
            continue
        ids = _validated_str_list(team.get("app_ids"), name, "app_ids")
        patterns = _validated_str_list(team.get("app_name_patterns"), name, "app_name_patterns")
        if ids is None or patterns is None:
            continue
        if not ids and not patterns:
            log.warning("Team '%s' has no app_ids or app_name_patterns and will match no apps", name)
        seen_names.add(name)
        valid.append({**team, "app_ids": ids, "app_name_patterns": [p.lower() for p in patterns]})
    return valid


def route_alerts(alerts: list[dict], teams: list[dict]) -> dict[str, list[dict]]:
    """Match alerts to each team by exact app ID or name pattern."""
    routed = {}
    for team in teams:
        ids = {i.lower() for i in (team.get("app_ids") or [])}
        patterns = list(team.get("app_name_patterns") or [])
        routed[team["name"]] = [
            a for a in alerts
            if a["app_id"].lower() in ids
            or any(p in a["app_name"].lower() for p in patterns)
        ]
    return routed


# ---------------------------------------------------------------------------
# Urgency buckets (derived from ALERT_THRESHOLD_DAYS)
# ---------------------------------------------------------------------------

def _bucket_labels(thresholds: list[int]) -> dict[str, str]:
    """Ordered mapping of bucket key -> human label."""
    labels = {"expired": "Expired"}
    for t in sorted(thresholds):
        labels[f"{t}_days"] = f"Within {t} days"
    labels[f"{max(thresholds)}_plus"] = f"{max(thresholds)}+ days"
    return labels


def _bucket_key(days_left: int, thresholds: list[int]) -> str:
    if days_left < 0:
        return "expired"
    for t in sorted(thresholds):
        if days_left <= t:
            return f"{t}_days"
    return f"{max(thresholds)}_plus"


def _bucket_counts(alerts: list[dict], thresholds: list[int]) -> dict:
    buckets = {key: 0 for key in _bucket_labels(thresholds)}
    for a in alerts:
        buckets[_bucket_key(a["days_left"], thresholds)] += 1
    return buckets


def _severity_color(key: str) -> str:
    """HTML colors matching the Slack bucket marks."""
    if key == "expired":
        return "#a71d2a"
    if key.endswith("_plus"):
        return "#0d6efd"
    t = int(key.split("_")[0])
    if t <= 7:
        return "#dc3545"
    if t <= 14:
        return "#6f42c1"
    if t <= 30:
        return "#d4a600"
    if t <= 60:
        return "#fd7e14"
    return "#0d6efd"


# ---------------------------------------------------------------------------
# State tracking - suppress duplicate notifications per channel
# ---------------------------------------------------------------------------

def _cred_key(a: dict) -> str:
    # keyId is unique per credential in Graph; fall back to display name for
    # credentials where it is missing.
    return f"{a['app_id']}|{a['credential_type']}|{a.get('key_id') or a['credential_name']}"


def _legacy_cred_key(a: dict) -> str:
    # Format used by state files written before key_id existed.
    return f"{a['app_id']}|{a['credential_type']}|{a['credential_name']}"


def _load_state() -> dict:
    """Return {channel_name: {"buckets": ..., "credential_keys": [...]}}."""
    if not STATE_FILE.exists():
        return {}
    try:
        data = json.loads(STATE_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}
    if not isinstance(data, dict):
        return {}
    channels = data.get("channels")
    if isinstance(channels, dict):
        return {k: v for k, v in channels.items() if isinstance(v, dict)}
    if "buckets" in data:
        # v1 single-channel format
        return {DEFAULT_CHANNEL: {
            "buckets": data.get("buckets", {}),
            "credential_keys": data.get("credential_keys", []),
        }}
    return {}


def _state_slice(buckets: dict, alerts: list[dict], transports: list[str] | None = None) -> dict:
    state = {
        "buckets": buckets,
        "credential_keys": sorted({_cred_key(a) for a in alerts}),
    }
    if transports is not None:
        state["transports"] = transports
    return state


def _save_state(channel_states: dict) -> None:
    state = {
        "version": 2,
        "last_run": datetime.now(timezone.utc).isoformat(),
        "channels": channel_states,
    }
    try:
        STATE_FILE.write_text(json.dumps(state, indent=2), encoding="utf-8")
    except OSError:
        log.exception("Could not write state file %s", STATE_FILE)


def _has_changes(current_buckets: dict, current_alerts: list[dict],
                 prev_slice: dict | None, thresholds: list[int]) -> tuple[bool, list[str], list[dict]]:
    """Compare a channel's current alerts against its previous state.

    Returns (changed, change descriptions, newly flagged alerts)."""
    if prev_slice is None:
        return True, ["First run - no previous state"], current_alerts

    changes = []
    prev_buckets = prev_slice.get("buckets", {})
    labels = _bucket_labels(thresholds)

    for key in dict.fromkeys(list(labels) + list(prev_buckets)):
        prev = prev_buckets.get(key, 0)
        curr = current_buckets.get(key, 0)
        if curr != prev:
            diff = curr - prev
            direction = f"+{diff}" if diff > 0 else str(diff)
            changes.append(f"{labels.get(key, key)}: {prev} -> {curr} ({direction})")

    prev_keys = set(prev_slice.get("credential_keys", []))
    # A credential counts as previously-known if either its keyId-based key or
    # its legacy name-based key was recorded (avoids a false "new" storm when
    # upgrading from state files without keyIds).
    matched_prev = set()
    new_alerts = []
    for a in current_alerts:
        known = {_cred_key(a), _legacy_cred_key(a)} & prev_keys
        if known:
            matched_prev |= known
        else:
            new_alerts.append(a)
    removed_creds = prev_keys - matched_prev

    if new_alerts:
        changes.append(f"{len(new_alerts)} new credential(s) flagged")
    if removed_creds:
        changes.append(f"{len(removed_creds)} credential(s) resolved/removed")

    return len(changes) > 0, changes, new_alerts


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def _safe_cell(ws, row: int, col: int, value):
    """Write a cell defensively: strip characters openpyxl rejects and force
    text type so values starting with '=' are never stored as formulas."""
    if isinstance(value, str):
        value = _ILLEGAL_XLSX_RE.sub("", value)
        cell = ws.cell(row=row, column=col, value=value)
        cell.data_type = "s"
        return cell
    return ws.cell(row=row, column=col, value=value)


def _generate_excel(alerts: list[dict]) -> str:
    """Generate an Excel report and return the file path."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Credential Expiry Report"

    headers = ["App Name", "App ID", "Type", "Credential Name", "Created", "Expires", "Days Left", "Status"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    red_font = Font(color="CC0000", bold=True)
    amber_font = Font(color="CC7A00", bold=True)
    blue_font = Font(color="0066CC")

    for row, a in enumerate(alerts, 2):
        _safe_cell(ws, row, 1, a["app_name"])
        _safe_cell(ws, row, 2, a["app_id"])
        _safe_cell(ws, row, 3, a["credential_type"])
        _safe_cell(ws, row, 4, a["credential_name"])
        _safe_cell(ws, row, 5, a.get("created", ""))
        _safe_cell(ws, row, 6, a["expires"])
        _safe_cell(ws, row, 7, a["days_left"])

        if a["days_left"] < 0:
            status = "EXPIRED"
            font = red_font
        elif a["days_left"] <= 7:
            status = "Critical"
            font = red_font
        elif a["days_left"] <= 30:
            status = "Warning"
            font = amber_font
        else:
            status = "Notice"
            font = blue_font

        status_cell = _safe_cell(ws, row, 8, status)
        status_cell.font = font

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="credential_expiry_report_")
    os.close(fd)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Slack notification
# ---------------------------------------------------------------------------

def _mrkdwn_escape(text: str) -> str:
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _bucket_emoji(thresholds: list[int]) -> dict[str, str]:
    """Color indicator per urgency bucket - every bucket gets a distinct mark."""
    by_value = {7: "🔴", 14: "🟣", 30: "🟡", 60: "🟠", 90: "🔵"}
    fallback = ["🔴", "🟡", "🟠", "⚪"]
    emoji = {"expired": "⛔"}
    for i, t in enumerate(sorted(thresholds)):
        emoji[f"{t}_days"] = by_value.get(t, fallback[min(i, len(fallback) - 1)])
    emoji[f"{max(thresholds)}_plus"] = "🔵"
    return emoji


def _friendly_date(iso_date: str) -> str:
    try:
        dt = datetime.strptime(iso_date, "%Y-%m-%d")
        return f"{dt.strftime('%b')} {dt.day}, {dt.year}"
    except ValueError:
        return iso_date


def _format_alert_lines(alerts: list[dict], thresholds: list[int]) -> list[str]:
    """Render alerts as reader-friendly mrkdwn lines.

    Recently expired and upcoming credentials lead as a "needs action" list;
    anything expired more than 30 days ago is assumed stale and collapses to
    a compact name list (full detail stays in the Excel report/email)."""
    emoji = _bucket_emoji(thresholds)
    recent_expired = sorted((a for a in alerts if -30 <= a["days_left"] < 0),
                            key=lambda a: a["days_left"], reverse=True)
    upcoming = sorted((a for a in alerts if a["days_left"] >= 0), key=lambda a: a["days_left"])
    stale = sorted((a for a in alerts if a["days_left"] < -30),
                   key=lambda a: a["days_left"], reverse=True)

    lines = []
    if recent_expired or upcoming:
        lines.append("🚨 *Needs action*")
        for a in recent_expired:
            ago = "today" if a["days_left"] == -0 else f"*{-a['days_left']} day(s) ago*"
            created = f" — created {_friendly_date(a['created'])}" if a.get("created") else ""
            lines.append(
                f"⛔  *{_mrkdwn_escape(a['app_name'])}* — {_mrkdwn_escape(a['credential_name'])} "
                f"({a['credential_type']}) — expired {ago} ({_friendly_date(a['expires'])}){created}"
            )
        for a in upcoming:
            d = a["days_left"]
            when = "*today*" if d == 0 else f"in *{d} day(s)*"
            dot = emoji.get(_bucket_key(d, thresholds), "🔵")
            created = f" — created {_friendly_date(a['created'])}" if a.get("created") else ""
            lines.append(
                f"{dot}  *{_mrkdwn_escape(a['app_name'])}* — {_mrkdwn_escape(a['credential_name'])} "
                f"({a['credential_type']}) — expires {when} ({_friendly_date(a['expires'])}){created}"
            )

    if stale:
        if lines:
            lines.append("")
        lines.append(f"🗑️ *Expired more than 30 days ago ({len(stale)})* — "
                     "likely unused; consider deleting these from Azure:")
        # Compact flowing name list, one app once (with a count when it has
        # several dead credentials), wrapped into chunk-friendly lines.
        counts: dict[str, int] = {}
        for a in stale:
            counts[a["app_name"]] = counts.get(a["app_name"], 0) + 1
        names = [f"{_mrkdwn_escape(n)} ({c})" if c > 1 else _mrkdwn_escape(n)
                 for n, c in counts.items()]
        max_names = 15  # most recently expired first; keep the message scannable
        shown = names[:max_names]
        row = ""
        for name in shown:
            if row and len(row) + len(name) + 3 > 200:
                lines.append(row)
                row = ""
            row = f"{row}  ·  {name}" if row else name
        if row:
            lines.append(row)
        if len(names) > max_names:
            lines.append(f"…and {len(names) - max_names} more — full list in the Excel report")
    return lines


def _chunk_lines(lines: list[str], budget: int = SECTION_CHAR_BUDGET) -> list[str]:
    """Join lines into newline-separated chunks, each within the char budget."""
    chunks, current, current_len = [], [], 0
    for line in lines:
        line = line[:budget]
        if current and current_len + len(line) + 1 > budget:
            chunks.append("\n".join(current))
            current, current_len = [], 0
        current.append(line)
        current_len += len(line) + 1
    if current:
        chunks.append("\n".join(current))
    return chunks


def _post_slack_webhook(webhook_url: str, blocks: list[dict], fallback: str) -> bool:
    try:
        resp = requests.post(webhook_url, json={"text": fallback, "blocks": blocks}, timeout=15)
        resp.raise_for_status()
        return True
    except Exception:
        log.exception("Slack webhook delivery failed")
        return False


def send_slack_alert(webhook_url: str, title: str, alerts: list[dict], buckets: dict,
                     changes: list[str], new_alerts: list[dict], thresholds: list[int],
                     excel_path: str | None, bot_token: str, channel_id: str,
                     action_hint: str = "") -> bool:
    """Send a channel report; returns True only if the message (and the Excel
    upload, when configured) was delivered."""
    labels = _bucket_labels(thresholds)
    emoji = _bucket_emoji(thresholds)
    upload_planned = bool(bot_token and channel_id and excel_path)
    first_run = bool(changes) and changes[0].startswith("First run")

    summary_fields = [
        {"type": "mrkdwn", "text": f"{emoji[key]} {label}:  *{buckets[key]}*"}
        for key, label in labels.items() if buckets.get(key)
    ]
    summary_block = {
        "type": "section",
        "text": {"type": "mrkdwn", "text": f"*{len(alerts)} credential(s) flagged*"},
    }
    if summary_fields:
        summary_block["fields"] = summary_fields[:10]

    blocks = [
        {
            "type": "header",
            "text": {"type": "plain_text", "text": title[:150]},
        },
        summary_block,
    ]

    if first_run:
        blocks.append({
            "type": "context",
            "elements": [{"type": "mrkdwn", "text": "First run — baseline established. "
                          "Future messages arrive only when something changes."}],
        })
    elif changes:
        changes_text = "*Changes since last run:*\n" + "\n".join(f"•  {c}" for c in changes)
        blocks.append({"type": "divider"})
        blocks.append({
            "type": "section",
            "text": {"type": "mrkdwn", "text": changes_text[:SECTION_CHAR_BUDGET]},
        })

    # List the newly flagged credentials, or - when a change fired without new
    # credentials (e.g. a bucket transition) - the full flagged set, so the
    # message always identifies what is expiring.
    listing = new_alerts if new_alerts else alerts
    if listing:
        heading = (f"*New credentials flagged ({len(new_alerts)}):*"
                   if new_alerts and not first_run
                   else f"*Currently flagged ({len(alerts)}):*")
        chunks = _chunk_lines(_format_alert_lines(listing, thresholds))
        # Stay under Slack's 50-block cap: header/summary/changes/footer use ~7.
        room = MAX_BLOCKS - len(blocks) - 3
        shown_chunks = chunks[:room]
        blocks.append({"type": "divider"})
        blocks.append({
            "type": "section",
            "text": {"type": "mrkdwn", "text": heading},
        })
        for chunk in shown_chunks:
            blocks.append({
                "type": "section",
                "text": {"type": "mrkdwn", "text": chunk},
            })
        if len(shown_chunks) < len(chunks):
            trunc_note = ("List truncated - see the attached Excel report for the full list."
                          if upload_planned else
                          "List truncated - run 'python monitor.py --dry-run' for the full list.")
            blocks.append({
                "type": "context",
                "elements": [{"type": "mrkdwn", "text": trunc_note}],
            })

    if action_hint:
        blocks.append({"type": "divider"})
        blocks.append({
            "type": "section",
            "text": {"type": "mrkdwn",
                     "text": f"ℹ️ *How to update:* {_mrkdwn_escape(action_hint)}"[:SECTION_CHAR_BUDGET]},
        })

    blocks.append({"type": "divider"})
    footer = "Full Excel report attached" if upload_planned else "Azure App Secret Expiry Monitor"
    blocks.append({
        "type": "context",
        "elements": [{
            "type": "mrkdwn",
            "text": ("👉 Rotated one of these? Reply in this thread so the team knows it's handled.\n"
                     f"{footer} | {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"),
        }],
    })

    fallback = f"{title}: {len(alerts)} credential(s) flagged"
    if not _post_slack_webhook(webhook_url, blocks, fallback):
        return False
    log.info("Slack alert sent: %s", title)

    if upload_planned:
        if not _upload_slack_file(bot_token, channel_id, excel_path):
            # The message promised an attachment; count the delivery as failed
            # so the whole notification retries next run (at-least-once).
            return False
    elif bot_token and channel_id:
        log.warning("Excel report unavailable - skipping Slack file upload")
    else:
        log.info("No bot token/channel ID for this channel - skipping file upload")
    return True


def send_slack_all_clear(webhook_url: str, title: str, resolved_count: int) -> bool:
    text = (
        f"✅ *{title}*\nAll clear: no credentials are currently within the alert thresholds. "
        f"{resolved_count} previously flagged credential(s) are no longer in scope "
        "(rotated, removed, or configuration changed)."
    )
    blocks = [{"type": "section", "text": {"type": "mrkdwn", "text": text[:SECTION_CHAR_BUDGET]}}]
    if not _post_slack_webhook(webhook_url, blocks, f"{title}: all clear"):
        return False
    log.info("Slack all-clear sent: %s", title)
    return True


def _upload_slack_file(bot_token: str, channel_id: str, file_path: str) -> bool:
    """Upload a file to Slack using the bot token."""
    report_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"credential_expiry_report_{report_date}.xlsx"

    try:
        # Step 1: Get an upload URL
        resp = requests.post(
            "https://slack.com/api/files.getUploadURLExternal",
            headers={"Authorization": f"Bearer {bot_token}"},
            data={"filename": filename, "length": os.path.getsize(file_path)},
            timeout=15,
        )
        resp.raise_for_status()
        result = resp.json()
        if not result.get("ok"):
            log.error("Slack files.getUploadURLExternal failed: %s", result.get("error"))
            return False

        upload_url = result["upload_url"]
        file_id = result["file_id"]

        # Step 2: Upload the file content
        with open(file_path, "rb") as f:
            resp = requests.post(upload_url, files={"file": (filename, f)}, timeout=30)
            resp.raise_for_status()

        # Step 3: Complete the upload and share to channel
        resp = requests.post(
            "https://slack.com/api/files.completeUploadExternal",
            headers={
                "Authorization": f"Bearer {bot_token}",
                "Content-Type": "application/json",
            },
            json={
                "files": [{"id": file_id, "title": filename}],
                "channel_id": channel_id,
            },
            timeout=15,
        )
        resp.raise_for_status()
        result = resp.json()
        if result.get("ok"):
            log.info("Excel report uploaded to Slack channel")
            return True
        log.error("Slack files.completeUploadExternal failed: %s", result.get("error"))
        return False
    except Exception:
        log.exception("Slack file upload failed")
        return False


# ---------------------------------------------------------------------------
# Email notification
# ---------------------------------------------------------------------------

def _smtp_send(subject: str, html: str, excel_path: str | None) -> bool | None:
    """Send an email. Returns None when email is not configured (so callers
    can exclude it from delivery gating), True/False for actual attempts."""
    smtp_host = os.getenv("SMTP_HOST", "")
    if not smtp_host:
        log.warning("SMTP_HOST not set, skipping email notification")
        return None

    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USERNAME", "")
    smtp_pass = os.getenv("SMTP_PASSWORD", "")
    use_tls = os.getenv("SMTP_USE_TLS", "true").lower() == "true"
    email_from = os.getenv("EMAIL_FROM", smtp_user)
    email_to = [e.strip() for e in os.getenv("EMAIL_TO", "").split(",") if e.strip()]

    if not email_to:
        log.warning("EMAIL_TO not set, skipping email notification")
        return None

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = email_from
    msg["To"] = ", ".join(email_to)
    msg.attach(MIMEText(html, "html"))

    if excel_path:
        with open(excel_path, "rb") as f:
            attachment = MIMEApplication(f.read(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            report_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            attachment.add_header("Content-Disposition", "attachment",
                                  filename=f"credential_expiry_report_{report_date}.xlsx")
            msg.attach(attachment)

    try:
        smtp_cls = smtplib.SMTP_SSL if smtp_port == 465 else smtplib.SMTP
        with smtp_cls(smtp_host, smtp_port, timeout=30) as server:
            if use_tls and smtp_cls is smtplib.SMTP:
                server.starttls()
            if smtp_user:
                server.login(smtp_user, smtp_pass)
            refused = server.sendmail(email_from, email_to, msg.as_string())
    except Exception:
        log.exception("Failed to send email alert")
        return False

    if refused:
        log.error("SMTP refused recipient(s): %s - will retry next run", ", ".join(refused))
        return False

    log.info("Email alert sent to %s", ", ".join(email_to))
    return True


def _email_detail_table(alerts: list[dict]) -> str:
    rows = ""
    for a in alerts:
        status = "EXPIRED" if a["days_left"] < 0 else f"{a['days_left']} days"
        color = "#dc3545" if a["days_left"] <= 7 else "#ffc107" if a["days_left"] <= 30 else "#17a2b8"
        rows += (
            f"<tr>"
            f"<td style='padding:6px 12px;border:1px solid #ddd'>{_html_escape(a['app_name'])}</td>"
            f"<td style='padding:6px 12px;border:1px solid #ddd'><code>{_html_escape(a['app_id'])}</code></td>"
            f"<td style='padding:6px 12px;border:1px solid #ddd'>{a['credential_type']}</td>"
            f"<td style='padding:6px 12px;border:1px solid #ddd'>{_html_escape(a['credential_name'])}</td>"
            f"<td style='padding:6px 12px;border:1px solid #ddd'>{a.get('created', '') or '-'}</td>"
            f"<td style='padding:6px 12px;border:1px solid #ddd'>{a['expires']}</td>"
            f"<td style='padding:6px 12px;border:1px solid #ddd;color:{color};font-weight:bold'>{status}</td>"
            f"</tr>"
        )
    return (
        "<table style='border-collapse:collapse'>"
        "<tr style='background:#f8f9fa'>"
        "<th style='padding:6px 12px;border:1px solid #ddd;text-align:left'>App Name</th>"
        "<th style='padding:6px 12px;border:1px solid #ddd;text-align:left'>App ID</th>"
        "<th style='padding:6px 12px;border:1px solid #ddd;text-align:left'>Type</th>"
        "<th style='padding:6px 12px;border:1px solid #ddd;text-align:left'>Credential</th>"
        "<th style='padding:6px 12px;border:1px solid #ddd;text-align:left'>Created</th>"
        "<th style='padding:6px 12px;border:1px solid #ddd;text-align:left'>Expires</th>"
        "<th style='padding:6px 12px;border:1px solid #ddd;text-align:left'>Remaining</th>"
        f"</tr>{rows}</table>"
    )


def _html_escape(text: str) -> str:
    return (text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def send_email_alert(alerts: list[dict], buckets: dict, thresholds: list[int],
                     excel_path: str | None) -> bool | None:
    """Send the email report. Returns None if email is unconfigured."""
    subject = f"Azure Credential Expiry Report - {len(alerts)} credential(s)"

    labels = _bucket_labels(thresholds)
    summary_rows = ""
    for key, label in labels.items():
        count = buckets.get(key, 0)
        if count:
            color = _severity_color(key)
            summary_rows += (
                f"<tr>"
                f"<td style='padding:6px 12px;border:1px solid #ddd'>{label}</td>"
                f"<td style='padding:6px 12px;border:1px solid #ddd;color:{color};font-weight:bold;text-align:center'>{count}</td>"
                f"</tr>"
            )

    attachment_note = (
        "<p>See the attached Excel report for the full credential list.</p>"
        if excel_path else ""
    )
    html = (
        f"<h2>Azure Credential Expiry Report</h2>"
        f"<p>{len(alerts)} credential(s) flagged across all monitored app registrations.</p>"
        f"<table style='border-collapse:collapse;margin-bottom:16px'>"
        f"<tr style='background:#f8f9fa'>"
        f"<th style='padding:6px 12px;border:1px solid #ddd;text-align:left'>Status</th>"
        f"<th style='padding:6px 12px;border:1px solid #ddd;text-align:center'>Count</th>"
        f"</tr>{summary_rows}</table>"
        f"{_email_detail_table(alerts)}"
        f"{attachment_note}"
    )
    return _smtp_send(subject, html, excel_path)


def send_email_all_clear(resolved_count: int) -> bool | None:
    html = (
        "<h2>Azure Credential Expiry Report</h2>"
        "<p>All clear: no credentials are currently within the alert thresholds. "
        f"{resolved_count} previously flagged credential(s) are no longer in scope "
        "(rotated, removed, or configuration changed).</p>"
    )
    return _smtp_send("Azure Credential Expiry Report - All clear", html, None)


# ---------------------------------------------------------------------------
# Channel processing
# ---------------------------------------------------------------------------

def _generate_excel_safe(alerts: list[dict], excel_paths: list[str]) -> str | None:
    try:
        path = _generate_excel(alerts)
        excel_paths.append(path)
        return path
    except Exception:
        log.exception("Failed to generate Excel report - continuing without attachment")
        return None


def process_team_channel(team: dict, subset: list[dict], thresholds: list[int],
                         prev_slice: dict | None, bot_token: str,
                         excel_paths: list[str]) -> tuple[dict | None, bool]:
    """Handle one team channel.

    Returns (new state slice or None to keep the previous one,
             delivery_failed flag for the process exit code)."""
    name = team["name"]
    buckets = _bucket_counts(subset, thresholds)
    changed, changes, new_alerts = _has_changes(buckets, subset, prev_slice, thresholds)

    if not changed:
        log.info("[%s] No changes since last run - skipping", name)
        return None, False

    new_slice = _state_slice(buckets, subset)
    title = f"Azure Credential Expiry - {name}"

    if not subset:
        if prev_slice is None or not prev_slice.get("credential_keys"):
            # Nothing flagged before or now; just record the empty slice.
            return new_slice, False
        resolved = len(prev_slice.get("credential_keys", []))
        if send_slack_all_clear(team["slack_webhook_url"], title, resolved):
            return new_slice, False
        return None, True

    log.info("[%s] Changes detected: %s", name, "; ".join(changes))
    channel_id = team.get("slack_channel_id", "")
    excel_path = _generate_excel_safe(subset, excel_paths) if (bot_token and channel_id) else None
    delivered = send_slack_alert(
        team["slack_webhook_url"], title, subset, buckets, changes, new_alerts,
        thresholds, excel_path, bot_token, channel_id,
        action_hint=str(team.get("action_hint") or ""),
    )
    return (new_slice, False) if delivered else (None, True)


def process_default_channel(alerts: list[dict], thresholds: list[int],
                            prev_slice: dict | None, bot_token: str,
                            excel_paths: list[str]) -> tuple[dict | None, bool]:
    """Handle the env-var-configured catch-all channel (Slack + email)."""
    buckets = _bucket_counts(alerts, thresholds)
    changed, changes, new_alerts = _has_changes(buckets, alerts, prev_slice, thresholds)

    slack_enabled = os.getenv("SLACK_ENABLED", "false").lower() == "true"
    email_enabled = os.getenv("EMAIL_ENABLED", "false").lower() == "true"
    webhook_url = os.getenv("SLACK_WEBHOOK_URL", "")
    channel_id = os.getenv("SLACK_CHANNEL_ID", "")
    transports = sorted(t for t, on in (("slack", slack_enabled), ("email", email_enabled)) if on)

    # Re-baseline when the set of enabled transports changes (e.g. Slack was
    # just turned on) so the newly enabled channel gets a first report.
    prev_transports = prev_slice.get("transports") if prev_slice else None
    if prev_transports is not None and prev_transports != transports:
        if not changed:
            changed = True
        changes.append("Notification channel configuration changed")

    if not changed:
        log.info("No changes since last run - skipping notifications")
        return None, False

    new_slice = _state_slice(buckets, alerts, transports)
    title = "Azure Credential Expiry Report"

    if not alerts:
        if prev_slice is None or not prev_slice.get("credential_keys"):
            return new_slice, False
        resolved = len(prev_slice.get("credential_keys", []))
        attempts = []
        if slack_enabled:
            if webhook_url:
                attempts.append(send_slack_all_clear(webhook_url, title, resolved))
            else:
                log.warning("SLACK_ENABLED but SLACK_WEBHOOK_URL not set - cannot send all-clear")
        if email_enabled:
            result = send_email_all_clear(resolved)
            if result is not None:
                attempts.append(result)
        if not attempts:
            log.warning("All %d previously flagged credential(s) resolved but no notification "
                        "channel is usable - will retry", resolved)
            return None, True
        return (new_slice, False) if all(attempts) else (None, True)

    log.info("Changes detected: %s", "; ".join(changes))

    if not slack_enabled and not email_enabled:
        log.warning("No notification channels enabled. Set SLACK_ENABLED=true or "
                    "EMAIL_ENABLED=true in .env - state NOT saved, alerts will retry")
        return None, True

    excel_path = _generate_excel_safe(alerts, excel_paths)

    # Every usable transport must deliver before the state slice is updated;
    # a partial failure means the next run re-sends to both (at-least-once).
    # Transports that are enabled but missing static config are warned about
    # and excluded, so one bad config cannot silence the healthy channel.
    attempts = []
    if slack_enabled:
        if webhook_url:
            attempts.append(send_slack_alert(
                webhook_url, title, alerts, buckets, changes, new_alerts,
                thresholds, excel_path, bot_token, channel_id,
            ))
        else:
            log.warning("SLACK_ENABLED but SLACK_WEBHOOK_URL not set - Slack skipped")
    if email_enabled:
        result = send_email_alert(alerts, buckets, thresholds, excel_path)
        if result is not None:
            attempts.append(result)

    if not attempts:
        log.warning("No usable notification channel configured - state NOT saved, alerts will retry")
        return None, True
    return (new_slice, False) if all(attempts) else (None, True)


# ---------------------------------------------------------------------------
# Channel wiring test
# ---------------------------------------------------------------------------

def test_channels(teams: list[dict]) -> int:
    """Send a test message (and, where configured, a sample Excel upload) to
    every configured Slack channel. Returns a process exit code."""
    bot_token = os.getenv("SLACK_BOT_TOKEN", "")
    targets = []

    if os.getenv("SLACK_WEBHOOK_URL", ""):
        targets.append(("default (env-configured channel)",
                        os.getenv("SLACK_WEBHOOK_URL", ""),
                        os.getenv("SLACK_CHANNEL_ID", "")))
    for team in teams:
        targets.append((team["name"], team["slack_webhook_url"], team.get("slack_channel_id", "")))

    if not targets:
        log.error("No Slack channels configured (SLACK_WEBHOOK_URL or channels.json)")
        return 1

    sample = [{
        "app_name": "Sample App", "app_id": "00000000-0000-0000-0000-000000000000",
        "credential_type": "Secret", "credential_name": "sample-secret",
        "key_id": "test", "expires": "2099-01-01", "days_left": 5,
    }]

    failures = 0
    for name, webhook_url, channel_id in targets:
        blocks = [
            {"type": "header", "text": {"type": "plain_text", "text": "Channel wiring test"[:150]}},
            {"type": "section", "text": {"type": "mrkdwn", "text": (
                f"This is a test from the Azure credential expiry monitor for channel *{_mrkdwn_escape(name)}*.\n"
                "If you can read this, message routing works. No real credentials are included."
            )}},
        ]
        ok = _post_slack_webhook(webhook_url, blocks, f"Channel wiring test: {name}")
        log.info("[%s] message: %s", name, "OK" if ok else "FAILED")
        if not ok:
            failures += 1
            continue

        if bot_token and channel_id:
            excel_path = None
            try:
                excel_path = _generate_excel(sample)
                upload_ok = _upload_slack_file(bot_token, channel_id, excel_path)
                log.info("[%s] file upload: %s", name, "OK" if upload_ok else "FAILED")
                if not upload_ok:
                    failures += 1
            finally:
                if excel_path:
                    try:
                        os.unlink(excel_path)
                    except OSError:
                        pass
        else:
            log.info("[%s] file upload: skipped (no bot token/channel ID)", name)

    return 1 if failures else 0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Azure App Registration credential expiry monitor")
    parser.add_argument("--dry-run", action="store_true",
                        help="Fetch and show channel routing without sending or saving state")
    parser.add_argument("--test-channels", action="store_true",
                        help="Send a test message to every configured Slack channel and exit")
    args = parser.parse_args(argv)

    teams_cfg = load_channels()
    config_ok = teams_cfg is not None
    teams = teams_cfg or []

    if args.test_channels:
        return test_channels(teams) or (0 if config_ok else 1)

    thresholds = parse_thresholds()
    log.info("Starting expiry check (thresholds: %s days)", thresholds)

    token = get_access_token()
    apps = get_applications(token)
    log.info("Fetched %d app registration(s)", len(apps))

    apps = filter_applications(apps)
    alerts = check_expiry(apps, thresholds)
    routed = route_alerts(alerts, teams)

    if alerts:
        buckets = _bucket_counts(alerts, thresholds)
        labels = _bucket_labels(thresholds)
        log.info("Found %d expiring credential(s): %s", len(alerts),
                 ", ".join(f"{labels[k]}: {v}" for k, v in buckets.items() if v))
    else:
        log.info("No credentials expiring within threshold. All clear!")

    if args.dry_run:
        print(f"\nChannel routing (dry run) - {len(alerts)} credential(s) total\n")
        rows = [("default (env-configured channel)", alerts)] + [(t["name"], routed[t["name"]]) for t in teams]
        for name, subset in rows:
            print(f"  {name}: {len(subset)} credential(s)")
            for a in subset:
                status = "EXPIRED" if a["days_left"] < 0 else f"{a['days_left']}d left"
                created = a.get("created") or "-"
                print(f"    - {a['app_name']} | {a['credential_type']} | {a['credential_name']} | created {created} | expires {a['expires']} | {status}")
            print()
        print("Dry run: nothing sent, state not saved.")
        return 0 if config_ok else 1

    prev_states = _load_state()
    new_states = dict(prev_states)
    bot_token = os.getenv("SLACK_BOT_TOKEN", "")
    excel_paths: list[str] = []
    delivery_failed = False

    try:
        result, failed = process_default_channel(alerts, thresholds,
                                                 prev_states.get(DEFAULT_CHANNEL), bot_token, excel_paths)
        delivery_failed |= failed
        if result is not None:
            new_states[DEFAULT_CHANNEL] = result

        for team in teams:
            result, failed = process_team_channel(team, routed[team["name"]], thresholds,
                                                  prev_states.get(team["name"]), bot_token, excel_paths)
            delivery_failed |= failed
            if result is not None:
                new_states[team["name"]] = result

        # Drop state for channels no longer configured - but never on a run
        # where channels.json existed and could not be read.
        if config_ok:
            active = {DEFAULT_CHANNEL} | {t["name"] for t in teams}
            new_states = {k: v for k, v in new_states.items() if k in active}

        if new_states != prev_states:
            _save_state(new_states)
    finally:
        for path in excel_paths:
            try:
                os.unlink(path)
            except OSError:
                pass

    return 1 if (delivery_failed or not config_ok) else 0


if __name__ == "__main__":
    sys.exit(main())
